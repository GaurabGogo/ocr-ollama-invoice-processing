"""Export bill records to Excel via openpyxl."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from bills_ocr.database import BillRow, row_to_public


def bills_to_xlsx_bytes(rows: list[BillRow]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "bills"
    headers = [
        "id",
        "created_at",
        "original_filename",
        "date",
        "vendor",
        "amount",
        "quantity",
        "purpose",
        "category",
        "validation_warnings",
        "ocr_text",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for i, row in enumerate(rows, start=2):
        pub = row_to_public(row)
        s = pub.structured
        warnings = "; ".join(pub.validation_warnings)
        values = [
            pub.id,
            pub.created_at,
            pub.original_filename,
            s.date,
            s.vendor,
            s.amount,
            s.quantity,
            s.purpose,
            s.category.value,
            warnings,
            pub.ocr_text,
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=i, column=col, value=val)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
